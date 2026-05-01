import os
import json
import glob
import numpy as np
import matplotlib.pyplot as plt
from sklearn.metrics import f1_score
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

RESULTS_DIR = "results"
PLOTS_DIR   = "plots"

# display names for the result files
APPROACH_LABELS = {
    "a1":               "A1 Raw",
    "a2":               "A2 Fixed MFCC",
    "a3":               "A3 Event-Triggered",
    "a4_high":          "A4 High Res",
    "a4_medium":        "A4 Medium Res",
    "a4_low":           "A4 Low Res",
    "a5_dim16":         "A5 Embed (16)",
    "a5_dim32":         "A5 Embed (32)",
    "a5_dim64":         "A5 Embed (64)",
    "a5_dim128":        "A5 Embed (128)",
    "a1_mixed":         "A1 Raw (mixed)",
    "a2_mixed":         "A2 Fixed MFCC (mixed)",
    "a3_mixed":         "A3 Event-Triggered (mixed)",
    "a4_high_mixed":    "A4 High Res (mixed)",
    "a4_medium_mixed":  "A4 Medium Res (mixed)",
    "a4_low_mixed":     "A4 Low Res (mixed)",
    "a5_dim16_mixed":   "A5 Embed 16 (mixed)",
    "a5_dim32_mixed":   "A5 Embed 32 (mixed)",
    "a5_dim64_mixed":   "A5 Embed 64 (mixed)",
    "a5_dim128_mixed":  "A5 Embed 128 (mixed)",
}


def load_results(path):
    with open(path) as f:
        return json.load(f)


def compute_metrics(records):
    """
    Returns a dict with accuracy, macro F1, avg bytes per utterance,
    avg RTT, and transmission rate.

    VAD skips (predicted=None) on "silence" samples count as correct — the whole
    point of event-triggered transmission is suppressing non-speech. Skips on real
    keyword samples still count as wrong (missed detection).
    """
    true_labels, pred_labels = [], []
    bytes_list, rtt_list = [], []
    transmitted = 0

    for r in records:
        true_labels.append(r["true"])
        if r["predicted"] is not None:
            pred_labels.append(r["predicted"])
        elif r["true"] == "silence":
            pred_labels.append("silence")   # correct suppression
        else:
            pred_labels.append("_skipped")  # missed keyword
        bytes_list.append(r["bytes"])
        if r["transmitted"]:
            transmitted += 1
            if r["rtt"] is not None:
                rtt_list.append(r["rtt"])

    total = len(records)
    correct = sum(t == p for t, p in zip(true_labels, pred_labels))

    # exclude "_skipped" from F1 since it's not a real class
    filtered = [(t, p) for t, p in zip(true_labels, pred_labels) if p != "_skipped"]
    f1_true = [t for t, _ in filtered]
    f1_pred = [p for _, p in filtered]
    macro_f1 = f1_score(f1_true, f1_pred, average="macro", zero_division=0) if filtered else 0.0

    return {
        "accuracy":          correct / total,
        "macro_f1":          macro_f1,
        "avg_bytes":         np.mean(bytes_list),
        "avg_rtt_ms":        np.mean(rtt_list) * 1000 if rtt_list else None,
        "transmission_rate": transmitted / total,
        "n_samples":         total,
    }


def print_table(results):
    header = f"{'Approach':<25} {'Accuracy':>9} {'Macro F1':>10} {'Avg Bytes':>11} {'Avg RTT (ms)':>13} {'TX Rate':>8}"
    print("\n" + header)
    print("-" * len(header))

    for tag, metrics in sorted(results.items()):
        label   = APPROACH_LABELS.get(tag, tag)
        acc     = f"{metrics['accuracy']:.4f}"
        f1      = f"{metrics['macro_f1']:.4f}"
        byt     = f"{metrics['avg_bytes']:.0f}"
        rtt     = f"{metrics['avg_rtt_ms']:.1f}" if metrics["avg_rtt_ms"] is not None else "N/A"
        tx_rate = f"{metrics['transmission_rate']:.2f}"
        print(f"{label:<25} {acc:>9} {f1:>10} {byt:>11} {rtt:>13} {tx_rate:>8}")


def plot_pareto(results):
    os.makedirs(PLOTS_DIR, exist_ok=True)

    # --- accuracy vs bandwidth ---
    fig, ax = plt.subplots(figsize=(9, 5))

    for tag, metrics in results.items():
        label = APPROACH_LABELS.get(tag, tag)
        ax.scatter(metrics["avg_bytes"], metrics["accuracy"], s=80, zorder=3, label=label)

    ax.set_xlabel("Avg bytes per utterance")
    ax.set_ylabel("Accuracy")
    ax.set_title("Accuracy vs Bandwidth — all approaches")
    ax.legend(fontsize=7, loc="lower right", framealpha=0.9)
    ax.grid(True, linestyle="--", alpha=0.5)
    plt.tight_layout()
    plt.savefig(os.path.join(PLOTS_DIR, "accuracy_vs_bandwidth.png"), dpi=150)
    plt.close()
    print(f"Saved plots/accuracy_vs_bandwidth.png")

    # --- accuracy vs latency (only approaches that have RTT data) ---
    rtt_results = {k: v for k, v in results.items() if v["avg_rtt_ms"] is not None}
    if rtt_results:
        fig, ax = plt.subplots(figsize=(9, 5))
        for tag, metrics in rtt_results.items():
            label = APPROACH_LABELS.get(tag, tag)
            ax.scatter(metrics["avg_rtt_ms"], metrics["accuracy"], s=80, zorder=3, label=label)

        ax.set_xlabel("Avg round-trip latency (ms)")
        ax.set_ylabel("Accuracy")
        ax.set_title("Accuracy vs Latency — all approaches")
        ax.legend(fontsize=7, loc="lower right", framealpha=0.9)
        ax.grid(True, linestyle="--", alpha=0.5)
        plt.tight_layout()
        plt.savefig(os.path.join(PLOTS_DIR, "accuracy_vs_latency.png"), dpi=150)
        plt.close()
        print(f"Saved plots/accuracy_vs_latency.png")

    # --- combined Pareto curves ---
    fig, ax = plt.subplots(figsize=(9, 5))

    a4_tags = ["a4_low", "a4_medium", "a4_high"]
    a4_tags = [t for t in a4_tags if t in results]
    if a4_tags:
        a4_bytes = [results[t]["avg_bytes"] for t in a4_tags]
        a4_acc   = [results[t]["accuracy"]  for t in a4_tags]
        ax.plot(a4_bytes, a4_acc, "o-", label="A4 Dynamic MFCC")

    a5_tags = ["a5_dim16", "a5_dim32", "a5_dim64", "a5_dim128"]
    a5_tags = [t for t in a5_tags if t in results]
    if a5_tags:
        a5_bytes = [results[t]["avg_bytes"] for t in a5_tags]
        a5_acc   = [results[t]["accuracy"]  for t in a5_tags]
        ax.plot(a5_bytes, a5_acc, "s-", label="A5 Learned Embedding")

    for tag in ["a1", "a2", "a3", "a3_mixed"]:
        if tag in results:
            label = APPROACH_LABELS[tag]
            ax.scatter(results[tag]["avg_bytes"], results[tag]["accuracy"],
                       marker="^", s=100, zorder=3, label=label)

    ax.set_xlabel("Avg bytes per utterance")
    ax.set_ylabel("Accuracy")
    ax.set_title("Pareto Curves — Accuracy vs Bandwidth")
    ax.legend()
    ax.grid(True, linestyle="--", alpha=0.5)
    plt.tight_layout()
    plt.savefig(os.path.join(PLOTS_DIR, "pareto_curves.png"), dpi=150)
    plt.close()
    print(f"Saved plots/pareto_curves.png")

    # --- A3: keywords-only vs mixed ---
    a3_tags = ["a3", "a3_mixed"]
    a3_tags = [t for t in a3_tags if t in results]
    if len(a3_tags) >= 1:
        fig, axes = plt.subplots(1, 3, figsize=(13, 4))

        metrics_list = [results[t] for t in a3_tags]
        labels       = [APPROACH_LABELS[t] for t in a3_tags]
        x            = np.arange(len(a3_tags))

        axes[0].bar(x, [m["accuracy"]          for m in metrics_list], color=["steelblue", "darkorange"][:len(a3_tags)])
        axes[0].set_xticks(x); axes[0].set_xticklabels(labels, fontsize=8)
        axes[0].set_ylabel("Accuracy"); axes[0].set_title("Accuracy")
        axes[0].set_ylim(0, 1); axes[0].grid(axis="y", linestyle="--", alpha=0.5)

        axes[1].bar(x, [m["avg_bytes"]          for m in metrics_list], color=["steelblue", "darkorange"][:len(a3_tags)])
        axes[1].set_xticks(x); axes[1].set_xticklabels(labels, fontsize=8)
        axes[1].set_ylabel("Avg bytes"); axes[1].set_title("Avg Bytes per Utterance")
        axes[1].grid(axis="y", linestyle="--", alpha=0.5)

        axes[2].bar(x, [m["transmission_rate"]  for m in metrics_list], color=["steelblue", "darkorange"][:len(a3_tags)])
        axes[2].set_xticks(x); axes[2].set_xticklabels(labels, fontsize=8)
        axes[2].set_ylabel("TX Rate"); axes[2].set_title("Transmission Rate")
        axes[2].set_ylim(0, 1); axes[2].grid(axis="y", linestyle="--", alpha=0.5)

        plt.suptitle("A3 Event-Triggered — Keywords-only vs Mixed", fontsize=11)
        plt.tight_layout()
        plt.savefig(os.path.join(PLOTS_DIR, "a3_comparison.png"), dpi=150)
        plt.close()
        print(f"Saved plots/a3_comparison.png")

    # --- A4: accuracy and F1 across resolutions ---
    a4_all = ["a4_low", "a4_medium", "a4_high"]
    a4_present = [t for t in a4_all if t in results]
    if a4_present:
        fig, axes = plt.subplots(1, 3, figsize=(13, 4))

        metrics_list = [results[t] for t in a4_present]
        labels       = [APPROACH_LABELS[t] for t in a4_present]
        x            = np.arange(len(a4_present))
        colors       = ["#d62728", "#ff7f0e", "#2ca02c"][:len(a4_present)]

        axes[0].bar(x, [m["accuracy"]  for m in metrics_list], color=colors)
        axes[0].set_xticks(x); axes[0].set_xticklabels(labels, fontsize=8)
        axes[0].set_ylabel("Accuracy"); axes[0].set_title("Accuracy by Resolution")
        axes[0].set_ylim(0, 1); axes[0].grid(axis="y", linestyle="--", alpha=0.5)

        axes[1].bar(x, [m["macro_f1"] for m in metrics_list], color=colors)
        axes[1].set_xticks(x); axes[1].set_xticklabels(labels, fontsize=8)
        axes[1].set_ylabel("Macro F1"); axes[1].set_title("Macro F1 by Resolution")
        axes[1].set_ylim(0, 1); axes[1].grid(axis="y", linestyle="--", alpha=0.5)

        axes[2].plot([m["avg_bytes"] for m in metrics_list],
                     [m["accuracy"]  for m in metrics_list],
                     "o-", color="steelblue")
        for i, (m, lbl) in enumerate(zip(metrics_list, labels)):
            axes[2].annotate(lbl, (m["avg_bytes"], m["accuracy"]),
                             textcoords="offset points", xytext=(6, 4), fontsize=8)
        axes[2].set_xlabel("Avg bytes"); axes[2].set_ylabel("Accuracy")
        axes[2].set_title("Accuracy vs Bandwidth"); axes[2].grid(True, linestyle="--", alpha=0.5)

        plt.suptitle("A4 Dynamic MFCC — Resolution Sweep", fontsize=11)
        plt.tight_layout()
        plt.savefig(os.path.join(PLOTS_DIR, "a4_resolution_sweep.png"), dpi=150)
        plt.close()
        print(f"Saved plots/a4_resolution_sweep.png")

    # --- A5: accuracy, F1, and RTT across embedding dims ---
    a5_all = ["a5_dim16", "a5_dim32", "a5_dim64", "a5_dim128"]
    a5_present = [t for t in a5_all if t in results]
    if a5_present:
        fig, axes = plt.subplots(1, 3, figsize=(13, 4))

        metrics_list = [results[t] for t in a5_present]
        labels       = [APPROACH_LABELS[t] for t in a5_present]
        dims         = [16, 32, 64, 128][:len(a5_present)]

        axes[0].plot(dims, [m["accuracy"]  for m in metrics_list], "o-", color="steelblue")
        axes[0].set_xlabel("Embedding dim"); axes[0].set_ylabel("Accuracy")
        axes[0].set_title("Accuracy vs Embedding Dim")
        axes[0].set_xticks(dims); axes[0].grid(True, linestyle="--", alpha=0.5)

        axes[1].plot(dims, [m["macro_f1"] for m in metrics_list], "s-", color="darkorange")
        axes[1].set_xlabel("Embedding dim"); axes[1].set_ylabel("Macro F1")
        axes[1].set_title("Macro F1 vs Embedding Dim")
        axes[1].set_xticks(dims); axes[1].grid(True, linestyle="--", alpha=0.5)

        axes[2].plot([m["avg_bytes"] for m in metrics_list],
                     [m["accuracy"]  for m in metrics_list],
                     "o-", color="steelblue")
        for m, lbl in zip(metrics_list, labels):
            axes[2].annotate(lbl, (m["avg_bytes"], m["accuracy"]),
                             textcoords="offset points", xytext=(6, 4), fontsize=8)
        axes[2].set_xlabel("Avg bytes"); axes[2].set_ylabel("Accuracy")
        axes[2].set_title("Accuracy vs Bandwidth"); axes[2].grid(True, linestyle="--", alpha=0.5)

        plt.suptitle("A5 Learned Embedding — Dimension Sweep", fontsize=11)
        plt.tight_layout()
        plt.savefig(os.path.join(PLOTS_DIR, "a5_embedding_sweep.png"), dpi=150)
        plt.close()
        print(f"Saved plots/a5_embedding_sweep.png")


def plot_results_table(results):
    os.makedirs(PLOTS_DIR, exist_ok=True)

    row_order = [
        "a1", "a2",
        "a3", "a3_mixed",
        "a4_high", "a4_medium", "a4_low",
        "a5_dim16", "a5_dim32", "a5_dim64", "a5_dim128",
    ]
    rows = [t for t in row_order if t in results]

    col_headers = ["Approach", "Accuracy", "Macro F1", "Avg Bytes", "Avg RTT (ms)", "TX Rate"]
    table_data  = []
    for tag in rows:
        m = results[tag]
        rtt = f"{m['avg_rtt_ms']:.1f}" if m["avg_rtt_ms"] is not None else "N/A"
        table_data.append([
            APPROACH_LABELS.get(tag, tag),
            f"{m['accuracy']:.4f}",
            f"{m['macro_f1']:.4f}",
            f"{m['avg_bytes']:.0f}",
            rtt,
            f"{m['transmission_rate']:.2f}",
        ])

    fig, ax = plt.subplots(figsize=(13, 0.5 * len(rows) + 1.5))
    ax.axis("off")

    tbl = ax.table(
        cellText=table_data,
        colLabels=col_headers,
        loc="center",
        cellLoc="center",
    )
    tbl.auto_set_font_size(False)
    tbl.set_fontsize(10)
    tbl.scale(1, 1.6)

    # style header row
    for col in range(len(col_headers)):
        tbl[0, col].set_facecolor("#2c3e50")
        tbl[0, col].set_text_props(color="white", fontweight="bold")

    # alternating row colors + highlight approach column
    group_colors = {
        "a1": "#d6eaf8", "a2": "#d6eaf8",
        "a3": "#d5f5e3", "a3_mixed": "#d5f5e3",
        "a4_high": "#fdebd0", "a4_medium": "#fdebd0", "a4_low": "#fdebd0",
        "a5_dim16": "#f9ebea", "a5_dim32": "#f9ebea",
        "a5_dim64": "#f9ebea", "a5_dim128": "#f9ebea",
    }
    for row_idx, tag in enumerate(rows, start=1):
        color = group_colors.get(tag, "#ffffff")
        for col in range(len(col_headers)):
            tbl[row_idx, col].set_facecolor(color)

    plt.title("Bandwidth-Efficient Keyword Spotting — Results Summary", fontsize=12, fontweight="bold", pad=12)
    plt.tight_layout()
    plt.savefig(os.path.join(PLOTS_DIR, "results_table.png"), dpi=150, bbox_inches="tight")
    plt.close()
    print(f"Saved plots/results_table.png")


def export_excel(results):
    os.makedirs(PLOTS_DIR, exist_ok=True)

    row_order = [
        "a1", "a2",
        "a3", "a3_mixed",
        "a4_high", "a4_medium", "a4_low",
        "a5_dim16", "a5_dim32", "a5_dim64", "a5_dim128",
    ]
    rows = [t for t in row_order if t in results]

    wb = Workbook()
    ws = wb.active
    ws.title = "Results"

    headers = ["Approach", "Accuracy", "Macro F1", "Avg Bytes", "Avg RTT (ms)", "TX Rate"]
    header_fill = PatternFill("solid", fgColor="2C3E50")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    group_colors = {
        "a1": "D6EAF8", "a2": "D6EAF8",
        "a3": "D5F5E3", "a3_mixed": "D5F5E3",
        "a4_high": "FDEBD0", "a4_medium": "FDEBD0", "a4_low": "FDEBD0",
        "a5_dim16": "F9EBEA", "a5_dim32": "F9EBEA",
        "a5_dim64": "F9EBEA", "a5_dim128": "F9EBEA",
    }

    for row_idx, tag in enumerate(rows, start=2):
        m = results[tag]
        rtt = round(m["avg_rtt_ms"], 1) if m["avg_rtt_ms"] is not None else "N/A"
        row_data = [
            APPROACH_LABELS.get(tag, tag),
            round(m["accuracy"], 4),
            round(m["macro_f1"], 4),
            int(m["avg_bytes"]),
            rtt,
            round(m["transmission_rate"], 2),
        ]
        fill = PatternFill("solid", fgColor=group_colors.get(tag, "FFFFFF"))
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center" if col > 1 else "left")
            cell.border = border

    # auto-fit column widths
    for col in range(1, len(headers) + 1):
        max_len = max(
            len(str(ws.cell(row=r, column=col).value or ""))
            for r in range(1, len(rows) + 2)
        )
        ws.column_dimensions[get_column_letter(col)].width = max_len + 4

    out_path = "results.xlsx"
    wb.save(out_path)
    print(f"Saved {out_path}")


def main():
    json_files = glob.glob(os.path.join(RESULTS_DIR, "*.json"))
    if not json_files:
        print(f"No result files found in {RESULTS_DIR}/")
        print("Run device.py for each approach first.")
        return

    results = {}
    for path in json_files:
        tag = os.path.splitext(os.path.basename(path))[0]
        records = load_results(path)
        results[tag] = compute_metrics(records)
        print(f"Loaded {tag}: {len(records)} samples")

    print_table(results)
    plot_pareto(results)
    plot_results_table(results)
    export_excel(results)


if __name__ == "__main__":
    main()
